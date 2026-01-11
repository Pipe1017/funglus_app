# backend/app/routers/informes_router.py
# CORRECCIÓN: Agregar Query a los imports

from typing import List
from fastapi import APIRouter, Depends, HTTPException, status, Query  # <-- AGREGADO Query aquí
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from app.db import database, models
from app.crud import crud_informes
from app.schemas import informe_schemas

router = APIRouter(
    prefix="/informes",
    tags=["Informes"],
)

@router.get(
    "/resumen/{ciclo_id}",
    response_model=List[informe_schemas.InformeResumenRow],
    summary="Obtener el resumen de resultados para un ciclo"
)
def get_resumen_ciclo(ciclo_id: int, db: Session = Depends(database.get_db)):
    informe_data = crud_informes.get_informe_resumen_by_ciclo(db=db, ciclo_id=ciclo_id)
    return informe_data


@router.post(
    "/historico",
    response_model=informe_schemas.HistoricoResponse,
    summary="Obtener datos históricos para comparar entre ciclos"
)
def get_datos_historicos(
    request: informe_schemas.HistoricoRequest,
    db: Session = Depends(database.get_db)
):
    """
    Recibe una métrica y una lista de combinaciones de catálogos,
    y devuelve una serie de datos históricos a través de todos los ciclos
    para ser graficada.
    """
    if len(request.combinaciones) > 4:
         raise HTTPException(status_code=400, detail="Se permite un máximo de 4 combinaciones para comparar.")

    return crud_informes.get_informe_historico(db=db, request=request)


# ===== NUEVO ENDPOINT PARA EXPORTAR A EXCEL =====
@router.get(
    "/resumen/export/excel/{ciclo_id}",
    summary="Exportar informe resumen a Excel",
    description="Genera un archivo Excel con la matriz MATRIS consolidada para uno o múltiples ciclos"
)
def export_resumen_excel(
    ciclo_id: int,
    ciclos_adicionales: str = Query(None, description="IDs de ciclos adicionales separados por comas, ej: '2,3,4'"),
    db: Session = Depends(database.get_db)
):
    """
    Exporta la tabla MATRIS a Excel.
    """
    
    # Obtener lista de ciclos a procesar
    ciclos_ids = [ciclo_id]
    if ciclos_adicionales:
        try:
            adicionales = [int(id.strip()) for id in ciclos_adicionales.split(',')]
            ciclos_ids.extend(adicionales)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Los IDs de ciclos adicionales deben ser números separados por comas"
            )
    
    # Crear workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe Resumen"
    
    # Estilos
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Encabezados
    headers = [
        'Ciclo', 'Etapa', 'Muestra', 'Origen', 'Tipo',
        'Fecha Ingreso', 'Cantidad',
        'Humedad %', 'Cenizas %', 'N Total %', 'N Seca %', 
        'pH', 'FDR Prom'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Procesar cada ciclo
    current_row = 2
    for ciclo_id_actual in ciclos_ids:
        # Obtener datos del ciclo
        try:
            informe_data = crud_informes.get_informe_resumen_by_ciclo(db, ciclo_id_actual)
        except Exception as e:
            print(f"Error al procesar ciclo {ciclo_id_actual}: {e}")
            continue
        
        # Obtener nombre del ciclo - CORREGIDO: usar models.Ciclo
        ciclo_obj = db.query(models.Ciclo).filter(models.Ciclo.id == ciclo_id_actual).first()
        ciclo_nombre = ciclo_obj.nombre_ciclo if ciclo_obj else f"Ciclo {ciclo_id_actual}"
        
        # Escribir datos
        for row_data in informe_data:
            ws.cell(row=current_row, column=1, value=ciclo_nombre)
            ws.cell(row=current_row, column=2, value=row_data.etapa_nombre)
            ws.cell(row=current_row, column=3, value=row_data.muestra_nombre)
            ws.cell(row=current_row, column=4, value=row_data.origen_nombre)
            ws.cell(row=current_row, column=5, value=row_data.tipo_agregacion)  # CORREGIDO
            ws.cell(row=current_row, column=6, value=row_data.fecha_ingreso)
            ws.cell(row=current_row, column=7, value=row_data.secuencias_count)
            
            # Resultados calculados
            ws.cell(row=current_row, column=8, value=row_data.resultado_humedad_prom_porc)
            ws.cell(row=current_row, column=9, value=row_data.resultado_cenizas_porc)
            ws.cell(row=current_row, column=10, value=row_data.resultado_nitrogeno_total_porc)
            ws.cell(row=current_row, column=11, value=row_data.resultado_nitrogeno_seca_porc)
            ws.cell(row=current_row, column=12, value=row_data.resultado_ph_valor)
            ws.cell(row=current_row, column=13, value=row_data.resultado_fdr_prom_kgf)
            
            # Aplicar bordes y alineación
            for col in range(1, 14):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col >= 8:  # Columnas numéricas
                    cell.alignment = right_align
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
            
            current_row += 1
        
        # Agregar fila vacía entre ciclos si hay múltiples
        if len(ciclos_ids) > 1:
            current_row += 1
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 12,
        'F': 14, 'G': 10,
        'H': 11, 'I': 11, 'J': 11, 'K': 11, 'L': 8, 'M': 9
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Generar archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"informe_resumen_{ciclo_nombre.replace(' ', '_')}_{timestamp}.xlsx"
    
    # Retornar respuesta con el archivo
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )